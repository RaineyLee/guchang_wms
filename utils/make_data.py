import openpyxl

class Location():
    def __init__(self, arg_1) :
        self.file_name = arg_1

    def excel_data(self):
        wb = openpyxl.load_workbook(filename = self.file_name)
        ws = wb.active

        all_values = []
        for row in ws.iter_rows(min_row=2):
            if row[0].value != '':
                row_value = []
                for cell in row:
                    row_value.append(cell.value)
                all_values.append(row_value)
            else:
                break

        return all_values
    
class Barcode():
    def __init__(self, arg_1) :
        self.file_name = arg_1

    def excel_data(self):
        wb = openpyxl.load_workbook(filename = self.file_name)
        ws = wb.active

        all_values = []
        for row in ws.iter_rows(min_row=2):
            if row[0].value != '':
                row_value = []
                for cell in row:
                    row_value.append(cell.value)
                all_values.append(row_value)
            else:
                break

        return all_values

    def column_title(self):
        wb = openpyxl.load_workbook(filename = self.file_name)
        ws = wb.active

        column_list = []
        for row in ws.iter_rows(max_row=1):
            if row[0].value != '':
                for cell in row:
                    column_list.append(cell.value) 

        return column_list

class Saleslist():
    def __init__(self, arg_1) :
        self.file_name = arg_1

    def excel_data(self):
        wb = openpyxl.load_workbook(filename = self.file_name)
        ws = wb.active

        all_values = []
        for row in ws.iter_rows(min_row=2):
            if row[0].value != '':
                row_value = []
                for cell in row:
                    row_value.append(str(cell.value))
                all_values.append(row_value)
            else:
                break

        return all_values

    def column_title(self):
        wb = openpyxl.load_workbook(filename = self.file_name)
        ws = wb.active

        column_list = []
        for row in ws.iter_rows(max_row=1):
            if row[0].value != '':
                for cell in row:
                    column_list.append(cell.value) 

        return column_list

        