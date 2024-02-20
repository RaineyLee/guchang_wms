'''
1. 파일 열기 
excelFile = openpyxl.load_workbook('파일명')  
excelFile 은 내가 임의로 지정한 파일명임
 
2. sheet 의 list 얻기 
(구) excelFile.get_sheet_names()
(신) excslFile.sheetnames
 
(구) 를 써도 에러가 나지는 않지만 새로운걸 써주는게 좋을듯
 
3. 특정 sheet 불러오기 
(구) excelFile.get_sheet_by_name('시트명')
(신) excelFile['시트명']
 
이거는 구버전을 쓰면
DeprecationWarning: Call to deprecated function get_sheet_names (Use wb.sheetnames)
Warning 이 나오기 때문에 사용하지 무조건 신버전을 사용하도록 하자
'''

### 원본 파일을 북청 "EXCEL 다운로드"가 아닌 "배송지 다운로드" 파일을 사용함

import openpyxl
import sys
from PyQt5.QtWidgets import *
from PyQt5.QtCore import QDate
from PyQt5 import uic
import os
import sys

# 절대경로를 상대경로로 변경 하는 함수
def resource_path(relative_path):
    try:
        base_path = sys._MEIPASS
    except Exception:
        base_path = os.path.abspath(".")

    return os.path.join(base_path, relative_path)

#UI파일 연결
#단, UI파일은 Python 코드 파일과 같은 디렉토리에 위치해야한다.
windows_ui_loc = "C:./CJ_number_v1.2.ui"
windows_excel_loc = "C:\\바탕화면\\"
# mac_ui_loc = "/Volumes/GoogleDrive/내 드라이브/개발/구창/CJ_number.ui"
# mac_excel_loc = "/Volumes/GoogleDrive/내 드라이브/개발/구창/출고정보.xlsx"

main_window = uic.loadUiType(resource_path(windows_ui_loc))[0]

#화면을 띄우는데 사용되는 Class 선언
class WindowClass(QMainWindow, main_window) :

    def __init__(self):
        super().__init__()
        self.setupUi(self)
        self.slot()

        self.date_edit.setDate(QDate.currentDate())

    def slot(self):
        self.btn_search.clicked.connect(self.file_open)
        self.btn_confirm.clicked.connect(self.load_excel)
        self.date_edit.dateChanged.connect(self.set_date)
        self.btn_select.clicked.connect(self.deleteRows)
        self.btn_save.clicked.connect(self.final_data)

    def set_date(self):
        date = self.date_edit.date().toString("yyyy-MM-dd")
        self.txt_date.setText(date)

    # 파일 오픈 대화상자
    def file_open(self):
        dialog = QFileDialog(self)
        fname = dialog.getOpenFileName(parent=self, caption='Open file', directory=windows_excel_loc)

        if fname[0]:
            self.txt_filename.setText(fname[0])
        else:
            QMessageBox.about(self, 'Warning', '파일을 선택하지 않았습니다.')

    # 파일 저장 대화상자
    def file_save(self):
        dialog = QFileDialog(self)
        qurl  = dialog.getSaveFileName(parent=self, caption='Open file', directory=windows_excel_loc)
        
        url = qurl[0]

        try:
            return url
        except Exception as e:
            QMessageBox.about(self, 'Warning', e)

        

    # openpyxl을 이용한 엑셀 파일 오픈
    # 셀값을 리스트로 만들어 한 라인만 만들기 위해 다른 함수로 보냄
    def load_excel(self):
        wb = openpyxl.load_workbook(filename = self.txt_filename.toPlainText())
        sheet = wb.active

        all_values = []
        for row in sheet.rows:
            if row[0].value != '':
                row_value = []
                for cell in row:
                    row_value.append(cell.value)
                all_values.append(row_value)
            else:
                break
        
        num = len(all_values)

        self.final_info(num, all_values) # 중복라인 제거용 함수로 전달

    # # load_excel함수에서 전달 받은 인자로 중복라인 제거하는 함수
    # def only_one_line(self, arr):
    #     second_values = []
    #     arr_all = arr

    #     for num in range(len(arr_all)-1):
    #         if arr_all[num][1] != arr_all[num+1][1]:
    #             second_values.append(arr_all[num+1])
    #         else:
    #             pass
        
    #     # self.main_info = second_values

    #     self.divide_address(second_values) # 주소 정보를 split하기 위한 함수로 전달

    # # only_one_line 함수에서 전달 받은 인자로 주소정보에서 '전화번호1, 전화번호2, 주소, 배송자(처)' 정보를 split
    # def divide_address(self, arr):
    #     add_values = []
    #     arr_all = arr

    #     for row in arr_all:
    #         var = row[11].split("/")
    #         add_values.append(var)

    #     # self.address_info = add_values
    #     self.final_info(len(arr_all), arr_all, add_values) # 원하는 형식의 엑셀파일을 만들기 위해 인자를 전달

    # 중복제거 된 배열과 주소정보만 추출한 정보를 합쳐서(텍스트 병합) 원하는 순서의 배열 생성
    def final_info(self, num, arr_1):
        com_arr = []
        fin_arr = []
        date = self.txt_date.toPlainText()

        for row in range(1, num):
            com_arr = ["", date, "사무용품", "1", arr_1[row][4], arr_1[row][5], arr_1[row][6], "", arr_1[row][7], arr_1[row][0]]
            fin_arr.append(com_arr)

            row_count = len(fin_arr)

        # 생성된 자료를 테이블로 만들어 눈으로 보고 확인 하게 함.
        self.make_table(row_count, fin_arr)
        
    # 테이블 생성 및 자료 표시
    def make_table(self, num, arr_1):
        self.tbl_list.setRowCount(num)
        col = self.tbl_list.columnCount()

        for i in range(num):
            for j in range(self.tbl_list.columnCount()): # 아니면 10개
                self.tbl_list.setItem(i, j, QTableWidgetItem(arr_1[i][j]))

        # 컨텐츠의 길이에 맞추어 컬럼의 길이를 자동으로 조절
        ################################################################
        table = self.tbl_list
        header = table.horizontalHeader()

        for i in range(col):
            header.setSectionResizeMode(i, QHeaderView.ResizeToContents)
        ################################################################
    
    # 테이블 선택범위 삭제
    def deleteRows(self):
        indexes = []
        rows = []

        for idx in self.tbl_list.selectedItems():
            indexes.append(idx.row())

        for value in indexes:
            if value not in rows:
                rows.append(value)

        # 삭제시 오류 방지를 위해 아래서 부터 삭제(리버스 소팅)
        rows = sorted(rows, reverse=True)

        # 선택행 삭제
        for rowid in rows:
            self.tbl_list.removeRow(rowid)

    
    # 테이블에 남겨진 정보를 엑셀로 변환
    def final_data(self):
        rows = self.tbl_list.rowCount()
        cols = self.tbl_list.columnCount()

        list_2 = [] # 최종적으로 사용할 리스트는 for문 밖에 선언

        for i in range(rows):
            list_1 = [] # 2번째 for문 안쪽에서 사용할 리스트 선언
            for j in range(cols):
                data = self.tbl_list.item(i,j)
                list_1.append(data.text())
            list_2.append(list_1)

        num = len(list_2)
        self.make_excel(list_2, num)
        

    # 엑셀 파일을 만들고 넘겨진 배열 정보를 이용하여 sheet에 정보를 기입/저장 함.
    def make_excel(self, arr, num):
        wb = openpyxl.Workbook()
        wb.create_sheet(index=0, title='CJ택배')

        sheet = wb.active
        list_line = ['번호', '날짜', '품목', '수량', '받는사람', '전화번호1', '전화번호2', '우편번호', '주소', '비고']
        sheet.append(list_line)

        for i in range(num):
            for j in range(len(list_line)):
                sheet.cell(row=i+2, column=j+1, value=arr[i][j])

        fname = self.file_save()

        if fname:
            self.save_excel(wb, fname)
        else:
            return


    def save_excel(self, workbook, file_name):
        workbook.save(file_name)

        self.close()




if __name__ == "__main__" :
    app = QApplication(sys.argv)
    myWindow = WindowClass()
    myWindow.show()
    app.exec()